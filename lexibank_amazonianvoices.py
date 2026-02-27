import pathlib
import re
import csv
import shutil
from typing import Optional
import datetime
import itertools
import unicodedata
from collections import defaultdict, OrderedDict
from collections.abc import Generator

import attr
import openpyxl
import pydub
import csvw
from csvw.metadata import URITemplate
import pylexibank
from clldutils.path import Path
from clldutils.misc import slug


@attr.s
class CustomConcept(pylexibank.Concept):
    Local_ID = attr.ib(default=None)
    Spanish_Gloss = attr.ib(default=None)
    Scientific_Name = attr.ib(default=None)
    Concepticon_SemanticField = attr.ib(default=None)


def fix_transcription(s):
    return s.replace('ɡ', 'g').replace(':', 'ː')


def norm(x):
    if x is None:
        return None
    return unicodedata.normalize('NFC', unicodedata.normalize('NFD', x))


def get_audio_id(d):
    aid = ''
    if 'audio_id' in d:
        aid = d['audio_id'].strip()
    elif 'audio_file' in d:
        aid = d['audio_file'].strip()
    if aid and aid != '-' and aid != '?':
        return aid
    return None


def get_concept_id(d, id_replacements):
    cid = d['concept-id'].replace('cbr_sr', 'cbrsr').strip()
    if cid == '':
        return None
    if cid in id_replacements:
        for np in id_replacements[cid]:
            if len(np[1]) == 0:
                cid = np[0]
            else:
                if d['spanish'] == np[1]:
                    cid = np[0]
                    break
    ids = cid.split('_')[1:]
    return '_'.join(map(lambda x: str(int(x)), ids))


def get_first_dir(p):
    for f in p.iterdir():
        if f.is_dir() and 'audio' in f.stem.lower():
            content = list(f.iterdir())
            if len(content) == 1 and content[0].is_dir():
                return content[0]
            return f
    return None


def get_audio_filename_via_slug(p, w):
    all_wavs = {}
    for wav in p.glob('**/*.wav'):
        wav_s = slug(wav.stem)
        if wav_s in all_wavs:
            return None
        all_wavs[wav_s] = wav.stem
    ws = slug(w)
    if ws in all_wavs:
        return all_wavs[ws]
    return None


class Dataset(pylexibank.Dataset):
    dir = Path(__file__).parent
    id = "amazonianvoices"

    concept_class = CustomConcept

    form_spec = pylexibank.FormSpec(
        missing_data=['-', '--', '- -'],
        replacements=[
            (':', 'ː'),
        ],
        normalize_unicode='NFC',
    )

    def cmd_download(self, args):
        """
        - converts ./raw/data/*.{wav,xlsx} into ./raw/csv/*.{lg_id/data.csv,lg_id/audio/*.wav}
        - all raw material (xlsx,wav) from Google Drive must be downloaded into ./raw/data first
        """
        valid_lg_ids = [lg['ID'] for lg in self.languages]
        valid_param_ids = [c['Local_ID'] for c in self.concepts]
        new_params_id_map = {c['Local_ID']: c['ID'] for c in self.concepts}
        # A duplicate concept - 199_several - snug into etc/concepts.csv. We must make sure it isn't
        # used for any word, so it doesn't make it into the CLDF data.
        fix_several_local_ids = {k for k, v in new_params_id_map.items() if v == '199_several'}
        for localid in fix_several_local_ids:
            new_params_id_map[localid] = '185_several'

        id_replacements = defaultdict(list)
        for row in self.etc_dir.read_csv('id_replacements.tsv', delimiter='\t'):
            id_replacements[row[1]].append((row[2], row[3]))

        datadir = self.raw_dir / 'data'
        for f in sorted(datadir.iterdir()):
            if f.is_dir() and f.glob('*.xlsx'):
                lg_id = slug(f.stem)
                if lg_id not in valid_lg_ids:
                    continue
                cdir = self.raw_dir / 'csv' / lg_id
                cdir.mkdir(exist_ok=True)
                try:
                    shutil.rmtree(self.raw_dir / 'csv' / lg_id / 'audio', ignore_errors=True)
                except Exception:
                    pass
                for xlsx in sorted(f.glob('**/*.xlsx')):
                    if 'concept' in xlsx.stem.lower() and '/~' not in str(xlsx):
                        wb = openpyxl.load_workbook(xlsx, data_only=True)
                        data = None
                        for sname in wb.sheetnames:
                            if sname.lower() == 'concepts':
                                data = list(self._iter_data(
                                    get_first_dir(f),
                                    lg_id,
                                    wb[sname],
                                    id_replacements,
                                    valid_param_ids,
                                    new_params_id_map,
                                    args.log,
                                ))
                                break
                        assert data, lg_id
                        with csvw.UnicodeWriter(cdir / 'data.csv') as w:
                            w.writerows(sorted(
                                data,
                                key=lambda i: float(i['param_id'].split('_')[0].replace('x', '.'))))
                        break
                else:
                    raise ValueError(f'No spreadsheet found in {f}')

    @staticmethod
    def _iter_rows(lg_id, sheet):
        rows = []
        for i, row in enumerate(sheet.rows):
            row = ['' if col.value is None else '{0}'.format(col.value).strip() for col in row]
            if i == 0:
                header = list(map(str.lower, row))
                continue
            assert header
            if 'english' not in header:
                raise ValueError(lg_id, header)
            d = dict(zip(header, row))
            word = d['segment'].replace('  ', ' ').strip()
            if word == '' or word in ['-', '--', '- -']:
                continue
            rows.append(d)

        for gloss, rr in itertools.groupby(
                sorted(rows, key=lambda r: r['english']), lambda r: r['english']):
            rr = list(rr)
            # If rows have the same english gloss and concept-ids with `_1`, `_2`, etc. suffixes,
            # we assume they all map to the same core concept.
            cids = [r['concept-id'] for r in rr]
            if len(set(cids)) != 1 and all(re.search(r'_[0-9]+_[123]$', cid) for cid in cids):
                for r in rr:
                    r['concept-id'] = '_'.join(r['concept-id'].split('_')[:-1])
            yield from rr

    def _iter_data(
            self,
            audio_dir,
            lg_id,
            sheet,
            id_replacements,
            valid_param_ids,
            new_params_id_map,
            log,
    ) -> Generator[OrderedDict[str, str], None, None]:
        all_words_for_pid = defaultdict(list)
        last_cid = None
        for d in self._iter_rows(lg_id, sheet):
            word = d['segment'].replace('  ', ' ').strip()
            cid = get_concept_id(d, id_replacements)
            if cid is None:
                if word and last_cid is not None:
                    cid = last_cid
                else:
                    log.warning(f'Ignoring word without cid {lg_id} {word}')
                    continue
            last_cid = cid
            if cid not in valid_param_ids:
                log.warning(f'{lg_id}\t{d["concept-id"]}\t{d["concept-id"]}')
                continue
            if word in all_words_for_pid[new_params_id_map[cid]]:
                log.warning(f'Ignoring duplicate word {lg_id} {cid} {word}')
                continue
            all_words_for_pid[new_params_id_map[cid]].append(word)
            aid = norm(get_audio_id(d))
            audio_name = ''
            if aid:
                aid = re.sub(r'\.(wav)?$', '', aid)
                audio_path = self._get_audio_path(audio_dir, aid, lg_id, cid, word, log)
                if audio_path:
                    audio_name = self._handle_audio(
                        audio_path, lg_id, cid, word, new_params_id_map, log)
                else:
                    log.warning(f'Audio ID but no audio for {lg_id} {cid} {word}')
            else:
                log.warning(f'No audio for {lg_id} {cid} {word}')
            yield OrderedDict(zip(
                ['param_id', 'form', 'audio'],
                [new_params_id_map[cid], fix_transcription(word), audio_name]))

    def _handle_audio(self, audio_path, lg_id, cid, word, new_params_id_map, log) -> str:
        ap = self.raw_dir / 'csv' / lg_id / 'audio'
        ap.mkdir(exist_ok=True)
        n = 1
        cid_ = new_params_id_map[cid]
        fp = ap / f'{lg_id}_{cid_}__{n}.wav'
        while fp.exists():
            n += 1
            fp = ap / f'{lg_id}_{cid_}__{n}.wav'

        audio_name = fp.stem

        wav = pydub.AudioSegment.from_file(str(audio_path), format='wav')
        if wav.channels > 1:
            wav_o = wav.split_to_mono()[0]
            if wav_o.rms < 10:
                wav_o = wav.split_to_mono()[1]
            if wav.rms < 10:
                log.warning(f'check audio for {lg_id} {cid} {word} "{audio_path}"')
        else:
            wav_o = wav
        wav_o = wav_o.fade_in(duration=50).fade_out(duration=50)
        wav_o = pydub.effects.normalize(wav_o, 4.)
        md = {
            'artist': lg_id,
            'title': '{}: {}'.format(audio_name, word),
            'album': 'amazonianvoices',
            'date': datetime.date.today().isoformat(),
            'genre': 'Speech'}
        wav_o.export(str(ap / audio_name) + '.wav', tags=md, format='wav', codec='copy')
        wav_o.export(str(ap / audio_name) + '.mp3', tags=md, format='mp3', bitrate='128k')
        wav_o.export(str(ap / audio_name) + '.ogg', tags=md, format='ogg', bitrate='128k')
        return audio_name

    def _get_audio_path(self, audio_dir, aid, lg_id, cid, word, log) -> Optional[pathlib.Path]:
        if '/' in aid:
            p, asid = aid.split('/')
            fp = audio_dir / p / f'{norm(asid.strip())}.wav'
            if fp.exists() and fp.is_file():
                return fp
        fp = audio_dir / f'{norm(aid)}.wav'
        if fp.exists() and fp.is_file():
            return fp
        n = get_audio_filename_via_slug(audio_dir, aid)
        fp = audio_dir / f'{n}.wav'
        if fp.exists() and fp.is_file():
            log.info(f'found heuristically {lg_id} {cid} {aid}')
            return fp
        log.warning(f'audio missing for {lg_id} {cid} {word} "{aid}"')
        return None

    def cmd_makecldf(self, args):
        with args.writer as ds:
            for lg in self.languages:
                ds.add_language(**lg)

            ds.cldf.add_component(
                'MediaTable',
                'objid',
                {'name': 'size', 'datatype': 'integer'},
                {
                    'name': 'Form_ID',
                    'required': True,
                    'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#formReference',
                    'datatype': 'string'
                },
                {
                    'name': 'mimetype',
                    'required': True,
                    'datatype': {'base': 'string', 'format': '[^/]+/.+'}
                },
            )
            ds.cldf.remove_columns('MediaTable', 'Download_URL')
            ds.cldf.remove_columns('MediaTable', 'Description')
            ds.cldf.remove_columns('MediaTable', 'Path_In_Zip')
            ds.cldf.remove_columns('MediaTable', 'Media_Type')
            ds.cldf['MediaTable', 'ID'].valueUrl = URITemplate('https://cdstar.eva.mpg.de/bitstreams/{objid}/{Name}')
            ds.cldf['MediaTable', 'mimetype'].propertyUrl = URITemplate('http://cldf.clld.org/v1.0/terms.rdf#mediaType')

            sound_cat = self.raw_dir.read_json('catalog.json')
            sound_map = dict()
            for k, v in sound_cat.items():
                sound_map[v['metadata']['name']] = k

            seen_param_ids = set()

            for lang_dir in pylexibank.progressbar(
                    sorted((self.raw_dir / 'csv').iterdir(), key=lambda f: f.name),
                    desc="adding new data"):

                if not lang_dir.is_dir():
                    continue

                lang_id = lang_dir.name

                with open(lang_dir / 'data.csv') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        new = ds.add_form(
                            Language_ID=lang_id,
                            Local_ID='',
                            Parameter_ID=row['param_id'],
                            Value=row['form'],
                            Form=self.form_spec.clean(row['form']),
                        )
                        seen_param_ids.add(row['param_id'])
                        if row['audio']:
                            media_id = row['audio']
                            if media_id in sound_map:
                                for bs in sorted(sound_cat[sound_map[media_id]]['bitstreams'],
                                                 key=lambda x: x['content-type']):
                                    ds.objects['MediaTable'].append({
                                        'ID': bs['checksum'],
                                        'Name': bs['bitstreamid'],
                                        'objid': sound_map[media_id],
                                        'mimetype': bs['content-type'],
                                        'size': bs['filesize'],
                                        'Form_ID': new['ID'],
                                    })
                            else:
                                args.log.warning(f'audio file {row["audio"]} not found in catalog')
                        else:
                            args.log.info(f'audio missing for {lang_id} {row["param_id"]}')

            for c in self.concepts:
                if c['ID'] in seen_param_ids:
                    ds.add_concept(**c)

            ds.objects['LanguageTable'].sort(key=lambda r: r['ID'])
            ds.objects['ParameterTable'].sort(key=lambda r: float(r['ID'].split('_')[0].replace('x', '.')))
            ds.objects['FormTable'].sort(key=lambda r: (
                r['Language_ID'],
                float(r['Parameter_ID'].split('_')[0].replace('x', '.')),
                int(r['ID'].split('-')[2])))
